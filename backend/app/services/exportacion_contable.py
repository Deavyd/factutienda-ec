from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.models.compra import Compra
from app.models.cuenta_cobrar import CuentaCobrar
from app.models.cuenta_pagar import CuentaPagar
from app.models.detalle_factura import DetalleFactura
from app.models.factura import Factura
from app.models.forma_pago_factura import FormaPagoFactura
from app.models.kardex import Kardex
from app.models.liquidacion_compra import LiquidacionCompra
from app.models.nota_credito import NotaCredito
from app.models.retencion import Retencion

MONEY_FMT = "$#,##0.00"
DATE_FMT = "DD/MM/YYYY"
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
ALT_FILL = PatternFill("solid", fgColor="F7F7F7")
SUM_FILL = PatternFill("solid", fgColor="E2F0D9")
RESUME_FILL = PatternFill("solid", fgColor="1F4E78")


def _d(value: object) -> Decimal:
    return Decimal(str(value or "0"))


def _style_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HDR_FILL


def _apply_table_style(ws, money_cols: set[int] | None = None, date_cols: set[int] | None = None) -> None:
    money_cols = money_cols or set()
    date_cols = date_cols or set()
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).fill = ALT_FILL
        for c in money_cols:
            ws.cell(row=row, column=c).number_format = MONEY_FMT
        for c in date_cols:
            ws.cell(row=row, column=c).number_format = DATE_FMT
    for cidx in range(1, ws.max_column + 1):
        max_len = 0
        col = ws.column_letter if False else None
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=cidx).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[ws.cell(row=1, column=cidx).column_letter].width = min(max(max_len + 2, 10), 42)


def _add_total_row(ws, values: list[object]) -> None:
    ws.append(values)
    r = ws.max_row
    for c in ws[r]:
        c.font = Font(bold=True)
        c.fill = SUM_FILL


def _factura_payment_map(db: Session, factura_ids: list[int]) -> dict[int, str]:
    if not factura_ids:
        return {}
    rows = db.query(FormaPagoFactura).filter(FormaPagoFactura.factura_id.in_(factura_ids)).all()
    out: dict[int, list[str]] = {}
    for r in rows:
        out.setdefault(r.factura_id, []).append(r.tipo)
    return {k: ", ".join(v) for k, v in out.items()}


def generar_reporte_contador(fecha_inicio: date, fecha_fin: date, db: Session) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    facturas = (
        db.query(Factura)
        .options(joinedload(Factura.cliente), joinedload(Factura.detalles))
        .filter(Factura.fecha_emision >= fecha_inicio, Factura.fecha_emision <= fecha_fin)
        .all()
    )
    factura_ids = [f.id for f in facturas]
    formas = _factura_payment_map(db, factura_ids)

    compras = db.query(Compra).options(joinedload(Compra.proveedor)).filter(Compra.fecha_emision >= fecha_inicio, Compra.fecha_emision <= fecha_fin).all()
    notas = db.query(NotaCredito).options(joinedload(NotaCredito.cliente), joinedload(NotaCredito.factura)).filter(NotaCredito.fecha_emision >= fecha_inicio, NotaCredito.fecha_emision <= fecha_fin).all()
    retenciones = db.query(Retencion).options(joinedload(Retencion.factura)).filter(Retencion.fecha_emision >= fecha_inicio, Retencion.fecha_emision <= fecha_fin).all()
    liquidaciones = db.query(LiquidacionCompra).filter(LiquidacionCompra.fecha_emision >= fecha_inicio, LiquidacionCompra.fecha_emision <= fecha_fin).all()
    cuentas_cobrar = db.query(CuentaCobrar).filter(CuentaCobrar.created_at >= datetime.combine(fecha_inicio, datetime.min.time()), CuentaCobrar.created_at <= datetime.combine(fecha_fin, datetime.max.time())).all()
    cuentas_pagar = db.query(CuentaPagar).filter(CuentaPagar.created_at >= datetime.combine(fecha_inicio, datetime.min.time()), CuentaPagar.created_at <= datetime.combine(fecha_fin, datetime.max.time())).all()

    total_ventas = sum((_d(f.total) for f in facturas), Decimal("0"))
    total_desc = sum((_d(f.descuento_total) for f in facturas), Decimal("0"))
    subtotal_0 = sum((_d(f.subtotal_0) for f in facturas), Decimal("0"))
    subtotal_15 = sum((_d(f.subtotal_12) for f in facturas), Decimal("0"))
    iva_15 = sum((_d(f.iva_total) for f in facturas), Decimal("0"))
    iva_5 = Decimal("0")
    autorizadas = sum(1 for f in facturas if (f.sri_estado or "").lower() == "autorizada")
    rechazadas = sum(1 for f in facturas if (f.sri_estado or "").lower() in {"devuelta", "rechazada", "error"})
    contingencia = sum(1 for f in facturas if (f.sri_estado or "").lower() == "contingencia")

    resumen_rows = [
        ("Periodo", f"{fecha_inicio} a {fecha_fin}"),
        ("Total ventas brutas", float(total_ventas)),
        ("Total descuentos", float(total_desc)),
        ("IVA cobrado 0%", 0.0),
        ("IVA cobrado 5%", float(iva_5)),
        ("IVA cobrado 15%", float(iva_15)),
        ("Total neto", float(total_ventas - total_desc)),
        ("Facturas emitidas", len(facturas)),
        ("Facturas autorizadas", autorizadas),
        ("Facturas rechazadas", rechazadas),
        ("Facturas contingencia", contingencia),
        ("Total compras", float(sum((_d(c.total) for c in compras), Decimal("0")))),
        ("Total retenciones recibidas", float(sum((_d(r.total_retenido) for r in retenciones), Decimal("0")))),
        ("RUC empresa", facturas[0].empresa.ruc if facturas else ""),
        ("Empresa", facturas[0].empresa.razon_social if facturas else ""),
    ]
    for idx, (k, v) in enumerate(resumen_rows, start=1):
        ws.cell(row=idx, column=1, value=k)
        ws.cell(row=idx, column=2, value=v)
    for c in ws[1]:
        c.fill = RESUME_FILL
        c.font = Font(color="FFFFFF", bold=True)
    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 32

    ws_f = wb.create_sheet("Facturas Emitidas")
    headers_f = ["Fecha emisión", "Número factura", "Clave de acceso", "Número autorización SRI", "RUC/Cédula cliente", "Nombre cliente", "Subtotal 0%", "Subtotal 5%", "Subtotal 15%", "Total descuento", "IVA 5%", "IVA 15%", "Total", "Forma de pago", "Estado SRI", "Observaciones"]
    _style_headers(ws_f, headers_f)
    for f in facturas:
        ws_f.append([
            f.fecha_emision,
            f.numero_comprobante,
            f.clave_acceso,
            f.sri_autorizacion,
            f.cliente.identificacion if f.cliente else "",
            f.cliente.razon_social if f.cliente else "",
            float(_d(f.subtotal_0)),
            0.0,
            float(_d(f.subtotal_12)),
            float(_d(f.descuento_total)),
            0.0,
            float(_d(f.iva_total)),
            float(_d(f.total)),
            formas.get(f.id, ""),
            f.sri_estado,
            "",
        ])
    _add_total_row(ws_f, ["TOTALES", "", "", "", "", "", float(subtotal_0), 0.0, float(subtotal_15), float(total_desc), 0.0, float(iva_15), float(total_ventas), "", "", ""])
    _apply_table_style(ws_f, money_cols={7, 8, 9, 10, 11, 12, 13}, date_cols={1})

    ws_nc = wb.create_sheet("Notas de Crédito")
    _style_headers(ws_nc, ["Fecha", "Número nota de crédito", "Número autorización", "Factura original referenciada", "Cliente", "Motivo", "Valor"])
    for n in notas:
        ws_nc.append([n.fecha_emision, n.numero_comprobante, n.clave_acceso, n.factura.numero_comprobante if n.factura else "", n.cliente.razon_social if n.cliente else "", n.motivo, float(_d(n.total))])
    _add_total_row(ws_nc, ["TOTALES", "", "", "", "", "", float(sum((_d(n.total) for n in notas), Decimal("0")))])
    _apply_table_style(ws_nc, money_cols={7}, date_cols={1})

    ws_c = wb.create_sheet("Compras")
    _style_headers(ws_c, ["Fecha", "Proveedor (nombre + RUC/cédula)", "Descripción", "Subtotal", "IVA", "Total", "Tipo documento", "Número documento"])
    for c in compras:
        ws_c.append([c.fecha_emision, f"{c.proveedor.razon_social} - {c.proveedor.identificacion}" if c.proveedor else "", "Compra", float(_d(c.subtotal)), float(_d(c.iva_total)), float(_d(c.total)), "factura", c.numero_documento])
    _add_total_row(ws_c, ["TOTALES", "", "", float(sum((_d(c.subtotal) for c in compras), Decimal("0"))), float(sum((_d(c.iva_total) for c in compras), Decimal("0"))), float(sum((_d(c.total) for c in compras), Decimal("0"))), "", ""])
    _apply_table_style(ws_c, money_cols={4, 5, 6}, date_cols={1})

    ws_r = wb.create_sheet("Retenciones Recibidas")
    _style_headers(ws_r, ["Fecha", "Agente de retención (nombre + RUC)", "Número retención", "Factura relacionada", "Base imponible renta", "% retención renta", "Valor retenido renta", "Base imponible IVA", "% retención IVA", "Valor retenido IVA", "Total retenido"])
    for r in retenciones:
        base_renta = pct_renta = val_renta = base_iva = pct_iva = val_iva = Decimal("0")
        for d in r.detalles or []:
            codigo = str(d.get("codigo", ""))
            base = _d(d.get("base_imponible"))
            pct = _d(d.get("porcentaje"))
            val = _d(d.get("valor"))
            if codigo.startswith("1"):
                base_renta += base
                val_renta += val
                pct_renta = pct
            else:
                base_iva += base
                val_iva += val
                pct_iva = pct
        ws_r.append([r.fecha_emision, f"{r.razon_social_agente} - {r.identificacion_agente}", r.numero_retencion, r.factura.numero_comprobante if r.factura else "", float(base_renta), float(pct_renta), float(val_renta), float(base_iva), float(pct_iva), float(val_iva), float(_d(r.total_retenido))])
    _add_total_row(ws_r, ["TOTALES", "", "", "", "", "", float(sum((_d(x.total_retenido) for x in retenciones), Decimal("0"))), "", "", "", float(sum((_d(x.total_retenido) for x in retenciones), Decimal("0")))])
    _apply_table_style(ws_r, money_cols={5, 7, 8, 10, 11}, date_cols={1})

    ws_l = wb.create_sheet("Liquidaciones de Compra")
    _style_headers(ws_l, ["Fecha", "Proveedor (nombre + cédula)", "Número autorización SRI", "Descripción", "Base 0%", "Base 15%", "IVA", "Total"])
    for l in liquidaciones:
        ws_l.append([l.fecha_emision, f"{l.proveedor_nombre} - {l.proveedor_cedula}", l.numero_autorizacion or "", "Liquidación de compra", float(_d(l.subtotal_0)), float(_d(l.subtotal_15)), float(_d(l.iva)), float(_d(l.total))])
    _add_total_row(ws_l, ["TOTALES", "", "", "", float(sum((_d(x.subtotal_0) for x in liquidaciones), Decimal("0"))), float(sum((_d(x.subtotal_15) for x in liquidaciones), Decimal("0"))), float(sum((_d(x.iva) for x in liquidaciones), Decimal("0"))), float(sum((_d(x.total) for x in liquidaciones), Decimal("0")))])
    _apply_table_style(ws_l, money_cols={5, 6, 7, 8}, date_cols={1})

    ws_i = wb.create_sheet("IVA Resumen Mensual")
    _style_headers(ws_i, ["Mes", "Ventas tarifa 0%", "Ventas tarifa 15%", "IVA cobrado", "Compras crédito tributario", "IVA pagado compras", "Retenciones IVA recibidas", "IVA a pagar / favor"])
    ventas_mensual = (
        db.query(func.strftime("%Y-%m", Factura.fecha_emision), func.sum(Factura.subtotal_0), func.sum(Factura.subtotal_12), func.sum(Factura.iva_total))
        .filter(Factura.fecha_emision >= fecha_inicio, Factura.fecha_emision <= fecha_fin)
        .group_by(func.strftime("%Y-%m", Factura.fecha_emision))
        .all()
    )
    compras_mensual = {row[0]: (_d(row[1]), _d(row[2])) for row in db.query(func.strftime("%Y-%m", Compra.fecha_emision), func.sum(Compra.subtotal), func.sum(Compra.iva_total)).filter(Compra.fecha_emision >= fecha_inicio, Compra.fecha_emision <= fecha_fin).group_by(func.strftime("%Y-%m", Compra.fecha_emision)).all()}
    for mes, v0, v15, iva_c in ventas_mensual:
        cbase, civa = compras_mensual.get(mes, (Decimal("0"), Decimal("0")))
        ret_iva = Decimal("0")
        iva_pagar = _d(iva_c) - civa - ret_iva
        ws_i.append([mes, float(_d(v0)), float(_d(v15)), float(_d(iva_c)), float(cbase), float(civa), float(ret_iva), float(iva_pagar)])
    _apply_table_style(ws_i, money_cols={2, 3, 4, 5, 6, 7, 8})

    ws_k = wb.create_sheet("Kardex Resumen")
    _style_headers(ws_k, ["Producto", "Stock inicial del período", "Total entradas (compras)", "Total salidas (ventas)", "Stock final", "Costo promedio", "Valor total inventario"])
    kardex_rows = (
        db.query(Kardex.producto_id)
        .filter(Kardex.created_at >= datetime.combine(fecha_inicio, datetime.min.time()), Kardex.created_at <= datetime.combine(fecha_fin, datetime.max.time()))
        .distinct()
        .all()
    )
    for (pid,) in kardex_rows:
        movs = db.query(Kardex).filter(Kardex.producto_id == pid).order_by(Kardex.created_at.asc()).all()
        if not movs:
            continue
        p = movs[0].producto
        periodo = [m for m in movs if fecha_inicio <= m.created_at.date() <= fecha_fin]
        if not periodo:
            continue
        stock_ini = _d(periodo[0].saldo_anterior)
        entradas = sum((_d(m.cantidad) for m in periodo if _d(m.cantidad) > 0), Decimal("0"))
        salidas = sum((abs(_d(m.cantidad)) for m in periodo if _d(m.cantidad) < 0), Decimal("0"))
        stock_fin = _d(periodo[-1].saldo_nuevo)
        costo_prom = sum((_d(m.costo_unitario) for m in periodo), Decimal("0")) / max(len(periodo), 1)
        ws_k.append([p.nombre if p else f"Producto {pid}", float(stock_ini), float(entradas), float(salidas), float(stock_fin), float(costo_prom), float(stock_fin * costo_prom)])
    _apply_table_style(ws_k, money_cols={6, 7})

    ws_cc = wb.create_sheet("Cuentas por Cobrar")
    _style_headers(ws_cc, ["Cliente", "Factura", "Fecha emisión", "Fecha vencimiento", "Monto total", "Monto pagado", "Saldo pendiente", "Estado"])
    for c in cuentas_cobrar:
        fac = db.query(Factura).options(joinedload(Factura.cliente)).filter(Factura.id == c.factura_id).first()
        ws_cc.append([
            fac.cliente.razon_social if fac and fac.cliente else "",
            fac.numero_comprobante if fac else "",
            fac.fecha_emision if fac else None,
            c.fecha_vencimiento,
            float(_d(c.monto_total)),
            float(_d(c.monto_pagado)),
            float(_d(c.monto_pendiente)),
            "VENCIDA" if c.fecha_vencimiento < date.today() and _d(c.monto_pendiente) > 0 else "VIGENTE",
        ])
    _add_total_row(ws_cc, ["TOTALES", "", "", "", float(sum((_d(x.monto_total) for x in cuentas_cobrar), Decimal("0"))), float(sum((_d(x.monto_pagado) for x in cuentas_cobrar), Decimal("0"))), float(sum((_d(x.monto_pendiente) for x in cuentas_cobrar), Decimal("0"))), ""])
    _apply_table_style(ws_cc, money_cols={5, 6, 7}, date_cols={3, 4})

    ws_cp = wb.create_sheet("Cuentas por Pagar")
    _style_headers(ws_cp, ["Proveedor", "Documento", "Fecha", "Fecha vencimiento", "Monto total", "Monto pagado", "Saldo pendiente", "Estado"])
    for c in cuentas_pagar:
        compra = db.query(Compra).options(joinedload(Compra.proveedor)).filter(Compra.id == c.compra_id).first()
        ws_cp.append([
            compra.proveedor.razon_social if compra and compra.proveedor else "",
            compra.numero_documento if compra else "",
            compra.fecha_emision if compra else None,
            c.fecha_vencimiento,
            float(_d(c.monto_total)),
            float(_d(c.monto_pagado)),
            float(_d(c.monto_pendiente)),
            "VENCIDA" if c.fecha_vencimiento < date.today() and _d(c.monto_pendiente) > 0 else "VIGENTE",
        ])
    _add_total_row(ws_cp, ["TOTALES", "", "", "", float(sum((_d(x.monto_total) for x in cuentas_pagar), Decimal("0"))), float(sum((_d(x.monto_pagado) for x in cuentas_pagar), Decimal("0"))), float(sum((_d(x.monto_pendiente) for x in cuentas_pagar), Decimal("0"))), ""])
    _apply_table_style(ws_cp, money_cols={5, 6, 7}, date_cols={3, 4})

    output = BytesIO()
    wb.save(output)

    ruc = facturas[0].empresa.ruc if facturas else "SINRUC"
    filename = f"FactuTienda_Contabilidad_{ruc}_{fecha_fin.month:02d}_{fecha_fin.year}.xlsx"
    return output.getvalue(), filename
