from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


def exportar_excel(datos: list, columnas: list, titulo: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws["A1"] = titulo
    ws["A2"] = f"Generado: {datetime.now().isoformat()}"
    for idx, col in enumerate(columnas, start=1):
        cell = ws.cell(row=4, column=idx, value=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    row = 5
    for item in datos:
        for idx, col in enumerate(columnas, start=1):
            ws.cell(row=row, column=idx, value=item.get(col))
        row += 1
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def exportar_pdf_reporte(datos: list, columnas: list, titulo: str, empresa: dict) -> bytes:
    out = io.BytesIO()
    c = canvas.Canvas(out, pagesize=A4)
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, 810, empresa.get("nombre", "FactuTienda EC"))
    c.setFont("Helvetica", 11)
    c.drawString(40, 792, titulo)
    y = 770
    c.setFont("Helvetica-Bold", 8)
    c.drawString(40, y, " | ".join(columnas))
    y -= 14
    c.setFont("Helvetica", 8)
    for item in datos[:45]:
        c.drawString(40, y, " | ".join(str(item.get(col, "")) for col in columnas)[:160])
        y -= 12
    c.drawString(40, 20, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')} - Pagina 1")
    c.save()
    return out.getvalue()
