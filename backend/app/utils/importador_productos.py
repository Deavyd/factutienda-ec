from __future__ import annotations

import base64
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


PLANTILLA_COLUMNS = [
    "codigo_principal",
    "codigo_auxiliar",
    "nombre",
    "descripcion",
    "precio_sin_iva",
    "precio_con_iva",
    "tarifa_iva",
    "unidad_venta",
    "unidad_compra",
    "factor_conversion",
    "stock_inicial",
    "stock_minimo",
    "tiene_ice",
    "categoria",
]

SRI_MAP = {
    "Código Principal": "codigo_principal",
    "Codigo Principal": "codigo_principal",
    "Código Auxiliar": "codigo_auxiliar",
    "Codigo Auxiliar": "codigo_auxiliar",
    "Descripción": "nombre",
    "Descripcion": "nombre",
    "Precio Unitario": "precio_sin_iva",
    "Tarifa IVA": "tarifa_iva",
    "Tiene ICE": "tiene_ice",
}


def _q2(value: object) -> Decimal:
    return Decimal(str(value or "0")).quantize(Decimal("0.01"), ROUND_HALF_UP)


def _to_bool_si_no(value: object) -> bool:
    return str(value or "NO").strip().upper() in {"SI", "S", "TRUE", "1", "YES"}


def _normalizar_row(row: dict, idx: int) -> tuple[dict | None, list[dict]]:
    errors: list[dict] = []
    codigo = str(row.get("codigo_principal", "")).strip()
    nombre = str(row.get("nombre", "")).strip().upper()
    if not codigo:
        errors.append({"fila": idx, "campo": "codigo_principal", "mensaje": "obligatorio"})
    if not nombre:
        errors.append({"fila": idx, "campo": "nombre", "mensaje": "obligatorio"})

    try:
        precio = _q2(row.get("precio_sin_iva", "0"))
        if precio <= 0:
            errors.append({"fila": idx, "campo": "precio_sin_iva", "mensaje": "debe ser > 0"})
    except Exception:
        errors.append({"fila": idx, "campo": "precio_sin_iva", "mensaje": "invalido"})
        precio = Decimal("0")

    tarifa_raw = str(row.get("tarifa_iva", "")).replace("%", "").strip()
    tarifa = tarifa_raw if tarifa_raw else "15"
    if tarifa not in {"0", "5", "15"}:
        errors.append({"fila": idx, "campo": "tarifa_iva", "mensaje": "solo 0, 5, 15"})

    unidad_venta = str(row.get("unidad_venta", "")).strip().lower()
    if not unidad_venta:
        errors.append({"fila": idx, "campo": "unidad_venta", "mensaje": "obligatoria"})

    unidad_compra = str(row.get("unidad_compra", "")).strip().lower() or unidad_venta
    factor_conversion = Decimal(str(row.get("factor_conversion", "1") or "1"))
    stock_inicial = Decimal(str(row.get("stock_inicial", "0") or "0"))
    stock_minimo = Decimal(str(row.get("stock_minimo", "0") or "0"))
    precio_con_iva = row.get("precio_con_iva")
    if precio_con_iva in (None, ""):
        precio_con_iva = _q2(precio * (Decimal("1") + Decimal(tarifa) / Decimal("100")))
    else:
        precio_con_iva = _q2(precio_con_iva)

    if errors:
        return None, errors

    return {
        "codigo_principal": codigo,
        "codigo_auxiliar": str(row.get("codigo_auxiliar", "") or "").strip() or None,
        "nombre": nombre,
        "descripcion": str(row.get("descripcion", "") or "").strip() or None,
        "precio_sin_iva": precio,
        "precio_con_iva": precio_con_iva,
        "tarifa_iva": Decimal(tarifa),
        "unidad_venta": unidad_venta,
        "unidad_compra": unidad_compra,
        "factor_conversion": factor_conversion,
        "stock_inicial": stock_inicial,
        "stock_minimo": stock_minimo,
        "tiene_ice": _to_bool_si_no(row.get("tiene_ice", "NO")),
        "categoria": str(row.get("categoria", "") or "").strip() or None,
    }, []


def _parse_excel_rows(archivo: bytes, map_columns: dict[str, str] | None = None) -> tuple[list[dict], list[dict]]:
    df = pd.read_excel(BytesIO(archivo))
    if map_columns:
        df = df.rename(columns=map_columns)
    registros: list[dict] = []
    errores: list[dict] = []
    for i, row in enumerate(df.to_dict(orient="records"), start=2):
        clean, errs = _normalizar_row(row, i)
        if errs:
            errores.extend(errs)
            continue
        registros.append(clean)
    return registros, errores


def generar_plantilla_importacion() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    ws.append(PLANTILLA_COLUMNS)
    ws.append(["P001", "AUX001", "ARROZ 1KG", "Arroz premium", 1.25, 1.44, 15, "und", "und", 1, 25, 5, "NO", "Granos"])
    for c in ws[1]:
        c.font = Font(bold=True)

    ins = wb.create_sheet("Instrucciones")
    ins.append(["Campo", "Regla"])
    ins.append(["codigo_principal", "Obligatorio y unico por empresa"])
    ins.append(["tarifa_iva", "Solo 0, 5 o 15"])
    ins.append(["unidad_venta", "Obligatoria. Si no existe, se crea"])
    ins.append(["precio_sin_iva", "Mayor a 0"])
    for c in ins[1]:
        c.font = Font(bold=True)

    u = wb.create_sheet("Unidades")
    u.append(["abreviatura", "nombre"])
    for val in ["und", "kg", "lb", "lt", "ml", "m", "cm", "caja", "paq"]:
        u.append([val, val.upper()])
    for c in u[1]:
        c.font = Font(bold=True)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def importar_desde_plantilla_sri(archivo: bytes) -> dict:
    registros, errores = _parse_excel_rows(archivo, SRI_MAP)
    total = len(registros) + len({e["fila"] for e in errores})
    return {
        "total": total,
        "importados": 0,
        "errores": errores,
        "duplicados": 0,
        "registros": registros,
    }


def importar_desde_excel_propio(archivo: bytes, mapeo_columnas: dict) -> dict:
    registros, errores = _parse_excel_rows(archivo, mapeo_columnas)
    total = len(registros) + len({e["fila"] for e in errores})
    if total > 0 and len({e["fila"] for e in errores}) / total > 0.5:
        return {
            "total": total,
            "importados": 0,
            "errores": errores,
            "duplicados": 0,
            "registros": [],
            "mensaje": "Mas del 50% de filas tienen errores. Corrija antes de importar.",
        }
    return {
        "total": total,
        "importados": 0,
        "errores": errores,
        "duplicados": 0,
        "registros": registros,
    }


def decode_base64_excel(archivo_base64: str) -> bytes:
    return base64.b64decode(archivo_base64)
